import os
import shutil
import uuid
from typing import Tuple
from fastapi import UploadFile, HTTPException
from pypdf import PdfReader
from openpyxl import load_workbook

# Base upload directory
UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "uploads"))

# Constraints
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
ALLOWED_EXTENSIONS = {".pdf", ".txt", ".xlsx"}

def get_workspace_dir(workspace_id: str) -> str:
    workspace_dir = os.path.join(UPLOAD_DIR, workspace_id)
    os.makedirs(workspace_dir, exist_ok=True)
    return workspace_dir

def validate_and_save_file(workspace_id: str, file: UploadFile) -> Tuple[str, str, int]:
    """
    Validates file extension and size, then saves it to the workspace directory.
    Returns (document_id, saved_file_path, file_size)
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Empty filename")
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")
        
    # Read file content for size checking
    content = file.file.read()
    file_size = len(content)
    
    if file_size == 0:
        raise HTTPException(status_code=400, detail="File is empty")
        
    if file_size > MAX_FILE_SIZE_BYTES:
        raise HTTPException(status_code=400, detail=f"File exceeds maximum size of 10MB")
        
    doc_id = str(uuid.uuid4())
    safe_filename = f"{doc_id}{ext}"
    workspace_dir = get_workspace_dir(workspace_id)
    file_path = os.path.join(workspace_dir, safe_filename)
    
    # Save the original file
    with open(file_path, "wb") as f:
        f.write(content)
        
    return doc_id, file_path, file_size

def extract_pdf_text(file_path: str) -> str:
    text = ""
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    except Exception as e:
        print(f"Error extracting PDF text: {e}")
    return text

def extract_xlsx_text(file_path: str) -> str:
    text = ""
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            text += f"--- Sheet: {sheetname} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join([str(cell) for cell in row if cell is not None])
                if row_str.strip():
                    text += row_str + "\n"
    except Exception as e:
        print(f"Error extracting XLSX text: {e}")
    return text

def extract_txt_text(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as e:
        print(f"Error extracting TXT text: {e}")
        return ""

def process_document(workspace_id: str, doc_id: str, file_path: str, original_filename: str) -> Tuple[int, str]:
    """
    Extracts text from the saved file, saves it to a .txt file, and returns (character_count, extracted_file_path)
    """
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    
    if ext == ".pdf":
        text = extract_pdf_text(file_path)
    elif ext == ".xlsx":
        text = extract_xlsx_text(file_path)
    elif ext == ".txt":
        text = extract_txt_text(file_path)
        
    # Normalize text
    text = "\n".join([line.strip() for line in text.split("\n") if line.strip()])
    
    # Save extracted text
    workspace_dir = get_workspace_dir(workspace_id)
    extracted_path = os.path.join(workspace_dir, f"{doc_id}_extracted.txt")
    
    with open(extracted_path, "w", encoding="utf-8") as f:
        f.write(text)
        
    return len(text), extracted_path

def delete_document_files(workspace_id: str, doc_id: str, ext: str):
    """
    Deletes the original file and the extracted text file.
    """
    workspace_dir = get_workspace_dir(workspace_id)
    original_file = os.path.join(workspace_dir, f"{doc_id}{ext}")
    extracted_file = os.path.join(workspace_dir, f"{doc_id}_extracted.txt")
    
    if os.path.exists(original_file):
        os.remove(original_file)
    if os.path.exists(extracted_file):
        os.remove(extracted_file)
